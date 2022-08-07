from openpyxl import load_workbook

load_wb = load_workbook('C:/Users/mye39/Desktop/mooho/AI/hwp_project/pyhwp/documents/오답노트/student_info/(22년)고1 오전클리닉 TEST.xlsx', data_only=True)

load_ws = load_wb['고1 여름방학 클리닉 TEST 6회']

print(load_ws['K6'].value)


#  # 지정한 셀의 값 출력

# get_cells = load_ws['k3' : 'k6']
# for row in get_cells:
#     for cell in row:
#         print(cell.value)

# 모든 행 단위로 출력

for i, row in enumerate(load_ws.rows):
    if i<10:
        val = row[2].value
        print(val)
        if val=='이름':
            print("location of name:", (i, 2))
    else:
        break
print()
# 모든 열 단위로 출력

# for column in load_ws.columns:
#     print(column)

# # # 모든 행과 열 출력

# all_values = []
# for row in load_ws.rows:
#     row_value = []
#     for cell in row:
#         row_value.append(cell.value)
#     all_values.append(row_value)
# print(all_values)

load_ws.cell(3, 3, 51470)
load_ws.cell(4, 3, 21470)
load_ws.cell(5, 3, 1470)
load_ws.cell(6, 3, 6470)

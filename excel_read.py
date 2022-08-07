import os
from openpyxl import load_workbook


def student_info_dict(file_location, sheet_number,student_info_dict=None):
    if student_info_dict==None:
        student_info_dict = {}
    
    load_wb = load_workbook(file_location, data_only=True)
    ws_names = load_wb.sheetnames
    # print("sheet names:",ws_names)
    load_ws = load_wb[ws_names[sheet_number-1]]
    
    
    # '이름'이라 적힌 칸의 위치를 찾는다.
    for i in range(1, 20):
        for j in range(1, 20):
            val = load_ws.cell(i, j).value
            if val=='이름':
                NameLocation = (i, j)
                Namerow, Namecol = NameLocation
                # print("Namelocation :", NameLocation)
                break 

    attendence_find = 0
    while True:
        attendence_find +=1
        if load_ws.cell(Namerow, Namecol+attendence_find).value=='출석':
            Attrow, Attcol = Namerow, Namecol+attendence_find
            # print('Attendence location :', Attrow, Attcol)
            break

    count = 0
    while True:
        count +=1
        row, col = Namerow+count, Namecol # 학생 이름의 위치. row, col
        student_name = load_ws.cell(row, col).value
        if student_name==None:
            break
        
        if (student_name != None) and (student_name not in student_info_dict):
            student_info_dict[student_name] = {}
        
        col_add = 0
        prob_list = []
        while True:
            if load_ws.cell(Namerow,Attcol+col_add).value==None:
                # print("Namerow:", Namerow)
                # print("prob_num:", Attcol+col_add)
                break
            col_add +=1
            if (load_ws.cell(row, Attcol+col_add).value != None) and (load_ws.cell(Namerow,Attcol+col_add).value!=None):
                prob_list.append(load_ws.cell(Namerow,Attcol+col_add).value)
        prob_list.reverse() # 복붙 방식이 먼저 붙여넣은게 뒤로 밀리는 방식이어서.. 문제 번호를 뒤집음.
        student_info_dict[student_name][sheet_number] = prob_list
        
    return student_info_dict
# file_location = 'C:/Users/mye39/Desktop/mooho/AI/hwp_project/pyhwp/documents/오답노트/student_info/(22년)고1 오전클리닉 TEST.xlsx'
# d = student_info_dict(file_location, 1)
# for key, val in d.items():
#     print("name:", key)
#     print("\tincorrect info:")
#     for testnum, probs in val.items():
#         print("\t\t", testnum, " : ",probs)
#     print()